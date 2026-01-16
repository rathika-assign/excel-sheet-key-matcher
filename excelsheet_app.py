import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Excel Sheet Matcher", layout="centered")
st.title("Excel Sheet Key Matcher (EmpID Only)")

uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])

if uploaded_file:
    xls = pd.ExcelFile(uploaded_file)
    sheet_names = xls.sheet_names

    main_sheet = st.selectbox("Select MAIN sheet", sheet_names)
    key_sheet = st.selectbox("Select KEY sheet", sheet_names)

    # ⚠️ HEADER IS ON ROW 2 (index 1)
    df_main = pd.read_excel(xls, main_sheet, header=1)
    df_key = pd.read_excel(xls, key_sheet, header=1)

    # Clean column names
    df_main.columns = df_main.columns.astype(str).str.strip()
    df_key.columns = df_key.columns.astype(str).str.strip()

    if "EmpID" not in df_main.columns or "EmpID" not in df_key.columns:
        st.error("❌ EmpID column must exist in both sheets (row 2 headers)")
    else:
        if st.button("Match & Highlight"):
            key_values = set(
                df_key["EmpID"]
                .dropna()
                .astype(str)
                .str.strip()
            )

            wb = load_workbook(uploaded_file)
            ws = wb[main_sheet]

            highlight = PatternFill(
                start_color="FFF59D",
                end_color="FFF59D",
                fill_type="solid"
            )

            # EmpID column index in Excel (still row 2 header)
            empid_col_index = list(df_main.columns).index("EmpID") + 1
            matched = 0

            # Data starts from ROW 3
            for row in range(3, ws.max_row + 1):
                value = ws.cell(row=row, column=empid_col_index).value

                if value is not None and str(value).strip() in key_values:
                    matched += 1
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = highlight

            wb.save("matched_highlighted.xlsx")

            st.success(f"✅ {matched} rows highlighted")

            st.download_button(
                "⬇ Download Highlighted Excel",
                data=open("matched_highlighted.xlsx", "rb"),
                file_name="matched_highlighted.xlsx"
            )
